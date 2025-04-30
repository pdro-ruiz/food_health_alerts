#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Script para la Generación Automática de Informes de Riesgo Alimentario

Este script genera tres tipos de informes a partir del análisis de riesgos alimentarios:
1. Un Excel segmentado por páginas para cada categoría de producto
2. Una presentación en PDF del notebook completo
3. Una presentación ejecutiva basada en los hallazgos principales

Uso:
    Este script está diseñado para ser importado como módulo en el pipeline principal
    o ejecutado directamente para generar informes específicos.
"""

import os
import argparse
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import re
import subprocess
import logging
import warnings
from matplotlib.patches import Patch

# Suprimir advertencias para una salida más limpia
warnings.filterwarnings('ignore')

# Configuración de logging
logger = logging.getLogger(__name__)

# Configuración global para visualizaciones
plt.style.use('ggplot')
sns.set(style="whitegrid")
plt.rcParams['figure.figsize'] = (10, 6)
plt.rcParams['font.size'] = 12

# Constantes para rutas de archivos
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
FINAL_DATA_DIR = os.path.join(DATA_DIR, 'final')
NOTEBOOKS_DIR = os.path.join(PROJECT_ROOT, 'notebooks')
REPORTS_DIR = os.path.join(PROJECT_ROOT, 'reports')

class AlertReportGenerator:
    """Clase para generar informes basados en el análisis de riesgos alimentarios"""
    
    def __init__(self, data_path=None, output_dir=None, notebook_path=None):
        """
        Inicializa el generador de informes
        
        Args:
            data_path (str): Ruta al archivo CSV de alertas alimentarias
            output_dir (str): Directorio donde se guardarán los informes
            notebook_path (str): Ruta al notebook de análisis
        """
        # Configurar rutas por defecto si no se especifican
        self.data_path = data_path or os.path.join(FINAL_DATA_DIR, 'consolidated_bakery_dairy_alerts.csv')
        self.output_dir = output_dir or REPORTS_DIR
        self.notebook_path = notebook_path or os.path.join(NOTEBOOKS_DIR, 'food_risk_analysis.ipynb')
        
        self.df = None
        self.report_date = datetime.now().strftime('%Y-%m-%d')
        
        # Crear directorios de salida si no existen
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(os.path.join(self.output_dir, 'excel'), exist_ok=True)
        os.makedirs(os.path.join(self.output_dir, 'pdf'), exist_ok=True)
        os.makedirs(os.path.join(self.output_dir, 'executive'), exist_ok=True)
        
        # Configurar etiquetas de severidad y probabilidad
        self.severidad_labels = {
            1: 'Menor',
            2: 'Moderada',
            3: 'Seria',
            4: 'Muy seria'
        }
        
        self.probabilidad_labels = {
            1: 'Remota',
            2: 'Ocasional',
            3: 'Probable',
            4: 'Frecuente'
        }
    
    def load_and_process_data(self):
        """Carga y procesa los datos, incluyendo la clasificación de riesgos"""
        logger.info("Cargando y procesando datos para generación de informes...")
        
        # Verificar que el archivo existe
        if not os.path.exists(self.data_path):
            logger.error(f"Archivo de datos no encontrado: {self.data_path}")
            return False
        
        try:
            # Cargar datos
            self.df = pd.read_csv(self.data_path)
            
            # Extraer año de la fecha
            self.df['year'] = self.df['date'].apply(self._extract_year)
            
            # Clasificar severidad
            self.df['severidad_num'] = self.df['hazard_type'].apply(self._clasificar_severidad)
            
            # Calcular frecuencias para probabilidad
            freq_origen = self._calcular_frecuencia_origen()
            freq_producto = self._calcular_frecuencia_producto()
            
            # Clasificar probabilidad
            self.df['probabilidad_num'] = self.df.apply(
                lambda row: self._clasificar_probabilidad(row, freq_origen, freq_producto), 
                axis=1
            )
            
            # Crear etiquetas para severidad y probabilidad
            self.df['severidad'] = self.df['severidad_num'].map(self.severidad_labels)
            self.df['probabilidad'] = self.df['probabilidad_num'].map(self.probabilidad_labels)
            
            # Calcular nivel de riesgo
            self.df['nivel_riesgo_num'] = self.df['severidad_num'] * self.df['probabilidad_num']
            
            # Clasificar nivel de riesgo
            self.df['nivel_riesgo'] = self.df['nivel_riesgo_num'].apply(self._clasificar_nivel_riesgo)
            
            logger.info(f"Datos procesados: {len(self.df)} alertas analizadas")
            return True
        
        except Exception as e:
            logger.error(f"Error al procesar datos: {str(e)}")
            return False
    
    def _extract_year(self, date_str):
        """Extrae el año de una cadena de fecha"""
        if pd.isna(date_str):
            return None
        
        match = re.search(r'\b(19|20)\d{2}\b', str(date_str))
        if match:
            return match.group(0)
        return None
    
    def _clasificar_severidad(self, hazard_type):
        """Clasifica la severidad del riesgo en escala 1-4 basado en el tipo de peligro"""
        if pd.isna(hazard_type):
            return 1
        
        hazard_lower = str(hazard_type).lower()
        
        # Peligros muy serios (4)
        if any(term in hazard_lower for term in [
            'listeria', 'salmonella', 'e. coli', 'escherichia coli', 'botulinum', 
            'aflatoxin', 'mercury', 'lead', 'cadmium', 'foreign body', 'metal', 
            'pieces', 'glass', 'cronobacter'
        ]):
            return 4
        
        # Peligros serios (3)
        elif any(term in hazard_lower for term in [
            'undeclared allergen', 'undeclared milk', 'undeclared peanut', 
            'undeclared soy', 'undeclared wheat', 'undeclared egg', 
            'undeclared nut', 'undeclared tree', 'pesticide', 'unauthorised substance',
            'moulds', 'high content', 'ethylene oxide'
        ]):
            return 3
        
        # Peligros moderados (2)
        elif any(term in hazard_lower for term in [
            'labelling', 'organoleptic', 'traces', 
            'too high count', 'presence', 'migration'
        ]):
            return 2
        
        # Por defecto
        else:
            return 1
    
    def _calcular_frecuencia_origen(self):
        """Calcula la frecuencia de aparición de cada país de origen"""
        return self.df['country_origin'].value_counts().to_dict()
    
    def _calcular_frecuencia_producto(self):
        """Calcula la frecuencia de aparición de cada tipo de producto"""
        return self.df['product_type'].value_counts().to_dict()
    
    def _clasificar_probabilidad(self, row, freq_origen, freq_producto):
        """Clasifica la probabilidad de ocurrencia en escala 1-4 basado en frecuencias"""
        origen = row['country_origin'] if not pd.isna(row['country_origin']) else "Desconocido"
        producto = row['product_type'] if not pd.isna(row['product_type']) else "Desconocido"
        
        # Obtener frecuencias
        freq_orig = freq_origen.get(origen, 0)
        freq_prod = freq_producto.get(producto, 0)
        
        # Normalizar frecuencias
        max_orig = max(freq_origen.values()) if freq_origen else 1
        max_prod = max(freq_producto.values()) if freq_producto else 1
        
        score_orig = freq_orig / max_orig if max_orig > 0 else 0
        score_prod = freq_prod / max_prod if max_prod > 0 else 0
        
        # Combinar scores (60% origen, 40% producto)
        combined_score = (score_orig * 0.6) + (score_prod * 0.4)
        
        # Clasificar
        if combined_score > 0.75:
            return 4  # Frecuente
        elif combined_score > 0.5:
            return 3  # Probable
        elif combined_score > 0.25:
            return 2  # Ocasional
        else:
            return 1  # Remota
    
    def _clasificar_nivel_riesgo(self, nivel):
        """Clasifica el nivel de riesgo basado en el producto de severidad y probabilidad"""
        if nivel <= 4:
            return 'Bajo'
        elif nivel <= 8:
            return 'Moderado'
        else:
            return 'Alto'

    def generate_excel_report(self):
        """
        Genera un informe Excel segmentado por páginas para cada categoría de producto
        
        Returns:
            str: Ruta al archivo Excel generado, o None si hubo un error
        """
        logger.info("Generando informe Excel...")
        
        # Importar solo si es necesario para no depender de openpyxl siempre
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference, PieChart, LineChart, Series
        except ImportError:
            logger.error("Error: Se requiere openpyxl para generar informes Excel.")
            logger.error("Instale con: pip install openpyxl")
            return None
        
        try:
            # Crear un nuevo libro de Excel
            wb = openpyxl.Workbook()
            
            # Eliminar la hoja por defecto
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Crear hoja de resumen
            self._create_summary_sheet(wb)
            
            # Crear hojas para cada categoría de producto
            categories = sorted(self.df['category'].dropna().unique())
            for category in categories:
                self._create_category_sheet(wb, category)
            
            # Guardar el archivo Excel
            excel_path = os.path.join(self.output_dir, 'excel', f'informe_riesgos_alimentarios_{self.report_date}.xlsx')
            wb.save(excel_path)
            
            logger.info(f"Informe Excel generado: {excel_path}")
            return excel_path
        
        except Exception as e:
            logger.error(f"Error al generar informe Excel: {str(e)}")
            return None
    
    def _create_summary_sheet(self, wb):
        """Crea la hoja de resumen en el libro Excel"""
        # Importar clases de openpyxl para asegurar que estén disponibles
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference, PieChart
        
        ws = wb.create_sheet("Resumen General")
        
        # Título
        ws['A1'] = "RESUMEN GENERAL DE ALERTAS ALIMENTARIAS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha del informe
        ws['A2'] = f"Fecha del informe: {self.report_date}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:H2')
        
        # Estadísticas generales - Títulos
        ws['A4'] = "Estadísticas Generales"
        ws['A4'].font = Font(bold=True)
        ws.merge_cells('A4:D4')
        
        # Estadísticas generales - Datos
        stats_row = 5
        ws[f'A{stats_row}'] = "Total de alertas:"
        ws[f'B{stats_row}'] = len(self.df)
        
        stats_row += 1
        ws[f'A{stats_row}'] = "Categorías de productos:"
        ws[f'B{stats_row}'] = len(self.df['category'].dropna().unique())
        
        stats_row += 1
        ws[f'A{stats_row}'] = "Países de origen:"
        ws[f'B{stats_row}'] = len(self.df['country_origin'].dropna().unique())
        
        stats_row += 1
        ws[f'A{stats_row}'] = "Tipos de peligros:"
        ws[f'B{stats_row}'] = len(self.df['hazard_type'].dropna().unique())
        
        # Distribución por nivel de riesgo
        stats_row += 2
        ws[f'A{stats_row}'] = "Distribución por Nivel de Riesgo"
        ws[f'A{stats_row}'].font = Font(bold=True)
        ws.merge_cells(f'A{stats_row}:D{stats_row}')
        
        risk_dist = self.df['nivel_riesgo'].value_counts()
        
        stats_row += 1
        headers = ["Nivel de Riesgo", "Cantidad", "Porcentaje", ""]
        for i, header in enumerate(headers):
            col = chr(65 + i)  # A, B, C, D
            ws[f'{col}{stats_row}'] = header
            ws[f'{col}{stats_row}'].font = Font(bold=True)
        
        for i, (nivel, count) in enumerate(risk_dist.items()):
            row = stats_row + i + 1
            ws[f'A{row}'] = nivel
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{count/len(self.df)*100:.1f}%"
        
        # Crear un gráfico de torta
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=stats_row+1, max_row=stats_row+len(risk_dist))
        data = Reference(ws, min_col=2, min_row=stats_row, max_row=stats_row+len(risk_dist))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribución por Nivel de Riesgo"
        
        # Añadir el gráfico de torta
        chart_row = stats_row + len(risk_dist) + 2
        ws.add_chart(pie, f"E{stats_row}")
        
        # Top países de origen
        ws[f'A{chart_row}'] = "Top 5 Países de Origen con Más Alertas"
        ws[f'A{chart_row}'].font = Font(bold=True)
        ws.merge_cells(f'A{chart_row}:D{chart_row}')
        
        top_countries = self.df['country_origin'].value_counts().head(5)
        
        chart_row += 1
        headers = ["País", "Cantidad", "", ""]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{chart_row}'] = header
            ws[f'{col}{chart_row}'].font = Font(bold=True)
        
        for i, (country, count) in enumerate(top_countries.items()):
            row = chart_row + i + 1
            ws[f'A{row}'] = country
            ws[f'B{row}'] = count
        
        # Crear un gráfico de barras para países
        bar = BarChart()
        data = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row+len(top_countries))
        cats = Reference(ws, min_col=1, min_row=chart_row+1, max_row=chart_row+len(top_countries))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.title = "Top 5 Países de Origen"
        
        # Añadir el gráfico de barras
        country_chart_row = chart_row + len(top_countries) + 2
        ws.add_chart(bar, f"E{chart_row}")
        
        # Top peligros
        ws[f'A{country_chart_row}'] = "Top 5 Tipos de Peligros Más Frecuentes"
        ws[f'A{country_chart_row}'].font = Font(bold=True)
        ws.merge_cells(f'A{country_chart_row}:D{country_chart_row}')
        
        top_hazards = self.df['hazard_type'].dropna().value_counts().head(5)
        
        hazard_row = country_chart_row + 1
        headers = ["Tipo de Peligro", "Cantidad", "", ""]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{hazard_row}'] = header
            ws[f'{col}{hazard_row}'].font = Font(bold=True)
        
        for i, (hazard, count) in enumerate(top_hazards.items()):
            row = hazard_row + i + 1
            hazard_text = str(hazard)
            if len(hazard_text) > 50:
                hazard_text = hazard_text[:47] + "..."
            ws[f'A{row}'] = hazard_text
            ws[f'B{row}'] = count
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        return ws
    
    def _create_category_sheet(self, wb, category):
        """Crea una hoja para una categoría específica de producto"""
        # Importar clases de openpyxl para asegurar que estén disponibles
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Filtrar datos para esta categoría
        df_cat = self.df[self.df['category'] == category].copy()
        
        # Si no hay datos, no crear la hoja
        if len(df_cat) == 0:
            return
        
        # Crear hoja
        sheet_name = category.capitalize()[:31]  # Excel limita nombres de hoja a 31 caracteres
        ws = wb.create_sheet(sheet_name)
        
        # Título
        ws['A1'] = f"ANÁLISIS DE ALERTAS: {category.upper()}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha del informe
        ws['A2'] = f"Fecha del informe: {self.report_date}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:H2')
        
        # Estadísticas de categoría
        ws['A4'] = "Estadísticas de la Categoría"
        ws['A4'].font = Font(bold=True)
        ws.merge_cells('A4:D4')
        
        stats_row = 5
        ws[f'A{stats_row}'] = "Total de alertas:"
        ws[f'B{stats_row}'] = len(df_cat)
        
        stats_row += 1
        ws[f'A{stats_row}'] = "Alertas de alto riesgo:"
        high_risk = len(df_cat[df_cat['nivel_riesgo'] == 'Alto'])
        ws[f'B{stats_row}'] = high_risk
        ws[f'C{stats_row}'] = f"{high_risk/len(df_cat)*100:.1f}%"
        
        stats_row += 1
        ws[f'A{stats_row}'] = "Principal tipo de peligro:"
        top_hazard = df_cat['hazard_type'].dropna().value_counts().index[0] if len(df_cat['hazard_type'].dropna()) > 0 else "N/A"
        top_hazard_text = str(top_hazard)
        if len(top_hazard_text) > 50:
            top_hazard_text = top_hazard_text[:47] + "..."
        ws[f'B{stats_row}'] = top_hazard_text
        
        stats_row += 1
        ws[f'A{stats_row}'] = "Principal país de origen:"
        top_country = df_cat['country_origin'].dropna().value_counts().index[0] if len(df_cat['country_origin'].dropna()) > 0 else "N/A"
        ws[f'B{stats_row}'] = top_country
        
        # Distribución por nivel de riesgo
        stats_row += 2
        ws[f'A{stats_row}'] = "Distribución por Nivel de Riesgo"
        ws[f'A{stats_row}'].font = Font(bold=True)
        ws.merge_cells(f'A{stats_row}:D{stats_row}')
        
        risk_dist = df_cat['nivel_riesgo'].value_counts()
        
        stats_row += 1
        headers = ["Nivel de Riesgo", "Cantidad", "Porcentaje", ""]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{stats_row}'] = header
            ws[f'{col}{stats_row}'].font = Font(bold=True)
        
        for i, (nivel, count) in enumerate(risk_dist.items()):
            row = stats_row + i + 1
            ws[f'A{row}'] = nivel
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{count/len(df_cat)*100:.1f}%"
        
        # Tendencia temporal
        trend_row = stats_row + len(risk_dist) + 2
        ws[f'A{trend_row}'] = "Evolución Temporal de Alertas"
        ws[f'A{trend_row}'].font = Font(bold=True)
        ws.merge_cells(f'A{trend_row}:D{trend_row}')
        
        # Calcular tendencia por año
        df_cat_year = df_cat[df_cat['year'].notna()].copy()
        year_counts = df_cat_year['year'].value_counts().sort_index()
        
        trend_row += 1
        headers = ["Año", "Cantidad", "", ""]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{trend_row}'] = header
            ws[f'{col}{trend_row}'].font = Font(bold=True)
        
        for i, (year, count) in enumerate(year_counts.items()):
            row = trend_row + i + 1
            ws[f'A{row}'] = year
            ws[f'B{row}'] = count
        
        # Top 5 países para esta categoría
        country_row = trend_row + len(year_counts) + 2
        ws[f'A{country_row}'] = "Top 5 Países de Origen"
        ws[f'A{country_row}'].font = Font(bold=True)
        ws.merge_cells(f'A{country_row}:D{country_row}')
        
        top_cat_countries = df_cat['country_origin'].dropna().value_counts().head(5)
        
        country_row += 1
        headers = ["País", "Cantidad", "Porcentaje", ""]
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{country_row}'] = header
            ws[f'{col}{country_row}'].font = Font(bold=True)
        
        for i, (country, count) in enumerate(top_cat_countries.items()):
            row = country_row + i + 1
            ws[f'A{row}'] = country
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{count/len(df_cat)*100:.1f}%"
        
        # Tabla de datos de alertas
        data_row = country_row + len(top_cat_countries) + 2
        ws[f'A{data_row}'] = "Detalles de Alertas"
        ws[f'A{data_row}'].font = Font(bold=True)
        ws.merge_cells(f'A{data_row}:H{data_row}')
        
        # Seleccionar columnas para mostrar
        df_display = df_cat[['date', 'product_name', 'hazard_type', 'country_origin', 
                             'severidad', 'probabilidad', 'nivel_riesgo']].copy()
        
        # Renombrar columnas para el Excel
        df_display.columns = ['Fecha', 'Producto', 'Tipo de Peligro', 'País de Origen', 
                              'Severidad', 'Probabilidad', 'Nivel de Riesgo']
        
        # Añadir los datos a la hoja
        data_row += 1
        for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws[f'{chr(65+c_idx)}{data_row+r_idx}']
                cell.value = value
                
                # Dar formato a los encabezados
                if r_idx == 0:
                    cell.font = Font(bold=True)
                
                # Colorear según nivel de riesgo
                if c_idx == 6 and r_idx > 0:  # Columna de nivel de riesgo, saltando encabezado
                    if value == 'Alto':
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif value == 'Moderado':
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    elif value == 'Bajo':
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        return ws

    def generate_notebook_pdf(self):
        """
        Convierte el notebook a PDF usando nbconvert
        
        Returns:
            str: Ruta al archivo PDF (o HTML) generado, o None si hubo un error
        """
        logger.info("Generando PDF del notebook...")
        
        # Verificar que existe
        if not os.path.exists(self.notebook_path):
            logger.error(f"Error: No se encuentra el notebook en {self.notebook_path}")
            return None
        
        try:
            # Convertir a HTML primero (es más confiable que directamente a PDF)
            html_output = os.path.join(self.output_dir, 'pdf', 'notebook_temp.html')
            cmd_html = [
                'jupyter', 'nbconvert', 
                '--to', 'html', 
                '--output', html_output,
                self.notebook_path
            ]
            
            subprocess.run(cmd_html, check=True)
            
            # Convertir HTML a PDF usando weasyprint
            try:
                from weasyprint import HTML
                pdf_path = os.path.join(self.output_dir, 'pdf', f'analisis_riesgos_alimentarios_{self.report_date}.pdf')
                HTML(html_output).write_pdf(pdf_path)
                logger.info(f"PDF del notebook generado: {pdf_path}")
                
                # Eliminar HTML temporal
                os.remove(html_output)
                
                return pdf_path
            
            except ImportError:
                logger.warning("WeasyPrint no está instalado. Se generará solo el HTML.")
                logger.warning("Instale con: pip install weasyprint")
                logger.info(f"HTML del notebook generado: {html_output}")
                return html_output
            
            except Exception as e:
                logger.error(f"Error al generar PDF con WeasyPrint: {str(e)}")
                logger.info(f"HTML del notebook generado como alternativa: {html_output}")
                return html_output
        
        except Exception as e:
            logger.error(f"Error al generar PDF: {str(e)}")
            return None

    def generate_executive_presentation(self):
        """
        Genera una presentación ejecutiva en PDF con los hallazgos principales
        
        Returns:
            str: Ruta al archivo PDF generado, o None si hubo un error
        """
        logger.info("Generando presentación ejecutiva...")
        
        try:
            from fpdf import FPDF
        except ImportError:
            logger.error("Error: Se requiere fpdf para generar la presentación ejecutiva.")
            logger.error("Instale con: pip install fpdf")
            return None
        
        try:
            # Crear directorio para figuras temporales
            temp_dir = os.path.join(self.output_dir, 'executive', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Generar figuras para la presentación
            self._generate_presentation_figures(temp_dir)
            
            # Establecer margen pequeño para mejor maquetación
            class PDF(FPDF):
                def header(self):
                    # Sin encabezado
                    pass
                def footer(self):
                    # Pie de página con número de página centrado
                    self.set_y(-15)
                    self.set_font('Arial', 'I', 8)
                    self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')
            
            # Crear PDF con márgenes reducidos
            pdf = PDF(orientation='L', unit='mm', format='A4')
            pdf.set_author("Sistema de Análisis de Riesgos Alimentarios")
            pdf.set_title("Resumen Ejecutivo: Análisis de Riesgos Alimentarios")
            pdf.set_margins(10, 10, 10)  # Margen izquierdo, superior, derecho
            pdf.set_auto_page_break(True, margin=15)  # Margen inferior
            
            # Portada
            pdf.add_page()
            pdf.set_font('Arial', 'B', 24)
            pdf.ln(30)  # Espacio desde el margen superior
            pdf.cell(0, 10, "ANÁLISIS DE RIESGOS ALIMENTARIOS", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font('Arial', 'B', 20)
            pdf.cell(0, 10, "Resumen Ejecutivo", ln=True, align='C')
            pdf.ln(30)
            pdf.set_font('Arial', '', 14)
            pdf.cell(0, 10, f"Fecha: {self.report_date}", ln=True, align='C')
            pdf.ln(50)
            pdf.set_font('Arial', 'I', 12)
            pdf.cell(0, 10, "Sistema de Análisis de Riesgos en Alertas Alimentarias", ln=True, align='C')
            pdf.cell(0, 10, "Productos de Panadería y Lácteos", ln=True, align='C')
            
            # Página de estadísticas principales
            pdf.add_page()
            pdf.set_font('Arial', 'B', 18)
            pdf.cell(0, 10, "Estadísticas Principales", ln=True)
            pdf.ln(5)
            pdf.set_font('Arial', '', 12)
            pdf.cell(0, 10, f"Total de alertas analizadas: {len(self.df)}", ln=True)
            
            # Distribución por nivel de riesgo
            risk_dist = self.df['nivel_riesgo'].value_counts()
            pdf.ln(5)
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, "Distribución por nivel de riesgo:", ln=True)
            pdf.set_font('Arial', '', 12)
            
            for nivel, count in risk_dist.items():
                pdf.cell(0, 10, f"- {nivel}: {count} alertas ({count/len(self.df)*100:.1f}%)", ln=True)
            
            # Añadir gráfico de distribución de riesgo
            pdf.ln(5)
            pdf.image(os.path.join(temp_dir, 'risk_distribution.png'), x=70, y=80, w=150)
            
            # Página de categorías con mayor riesgo
            pdf.add_page()
            pdf.set_font('Arial', 'B', 18)
            pdf.cell(0, 10, "Categorías con Mayor Riesgo", ln=True)
            pdf.ln(5)
            pdf.set_font('Arial', '', 12)
            pdf.multi_cell(0, 6, "El siguiente gráfico muestra la distribución de alertas por nivel de riesgo en cada categoría de producto. Esto permite identificar qué categorías presentan mayor concentración de alertas de alto riesgo.")
            
            # Añadir gráfico de categorías
            pdf.ln(5)
            pdf.image(os.path.join(temp_dir, 'risk_by_category.png'), x=15, y=50, w=260)
            
            pdf.ln(120)  # Espacio para la imagen
            pdf.set_font('Arial', '', 12)
            high_risk = self.df[self.df['nivel_riesgo'] == 'Alto']
            high_risk_cat = high_risk['category'].value_counts().head(3)
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, "Top 3 categorías con mayor número de alertas de alto riesgo:", ln=True)
            pdf.set_font('Arial', '', 12)
            
            for cat, count in high_risk_cat.items():
                pdf.cell(0, 10, f"- {cat}: {count} alertas", ln=True)
            
            # Página de países de origen
            pdf.add_page()
            pdf.set_font('Arial', 'B', 18)
            pdf.cell(0, 10, "Países de Origen con Mayor Incidencia", ln=True)
            pdf.ln(5)
            pdf.set_font('Arial', '', 12)
            pdf.multi_cell(0, 6, "Este análisis identifica los países de origen que presentan mayor concentración de alertas de alta severidad, lo que permite priorizar controles para importaciones de estas regiones.")
            
            # Añadir gráfico de países
            pdf.ln(5)
            pdf.image(os.path.join(temp_dir, 'top_countries.png'), x=15, y=50, w=260)
            
            pdf.ln(120)  # Espacio para la imagen
            top_countries = self.df['country_origin'].value_counts().head(5)
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, "Top 5 países de origen con más alertas:", ln=True)
            pdf.set_font('Arial', '', 12)
            
            for country, count in top_countries.items():
                pdf.cell(0, 10, f"- {country}: {count} alertas", ln=True)
            
            # Página de tendencias temporales
            pdf.add_page()
            pdf.set_font('Arial', 'B', 18)
            pdf.cell(0, 10, "Evolución Temporal de Alertas", ln=True)
            pdf.ln(5)
            pdf.set_font('Arial', '', 12)
            pdf.multi_cell(0, 6, "El análisis temporal muestra la evolución de alertas por nivel de riesgo a lo largo de los años, permitiendo identificar tendencias y evaluar el impacto de medidas de control.")
            
            # Añadir gráfico de tendencias
            pdf.ln(5)
            pdf.image(os.path.join(temp_dir, 'risk_trends.png'), x=15, y=50, w=260)
            
            pdf.ln(120)  # Espacio para la imagen
            pdf.set_font('Arial', '', 12)
            pdf.multi_cell(0, 6, "El análisis temporal muestra una evolución significativa en las alertas de riesgo a lo largo de los años. Se observa un incremento constante en alertas de alto riesgo, lo que sugiere una mayor detección o un aumento en los problemas de seguridad alimentaria.")
            
            # Página de principales peligros
            pdf.add_page()
            pdf.set_font('Arial', 'B', 18)
            pdf.cell(0, 10, "Tipos de Peligros Más Frecuentes", ln=True)
            pdf.ln(5)
            pdf.set_font('Arial', '', 12)
            pdf.multi_cell(0, 6, "Este análisis identifica los principales tipos de peligros detectados en las alertas, lo que permite enfocar los programas de control y prevención en los riesgos más relevantes.")
            
            # Añadir gráfico de peligros
            pdf.ln(5)
            pdf.image(os.path.join(temp_dir, 'top_hazards.png'), x=15, y=50, w=260)
            
            pdf.ln(120)  # Espacio para la imagen
            top_hazards_all = self.df['hazard_type'].dropna().value_counts().head(5)
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, "Top 5 tipos de peligros más frecuentes:", ln=True)
            pdf.set_font('Arial', '', 12)
            
            for hazard, count in top_hazards_all.items():
                hazard_text = str(hazard)
                if len(hazard_text) > 50:
                    hazard_text = hazard_text[:47] + "..."
                pdf.multi_cell(0, 8, f"- {hazard_text}: {count} alertas")
                pdf.ln(2)
            
            # Página de recomendaciones
            pdf.add_page()
            pdf.set_font('Arial', 'B', 18)
            pdf.cell(0, 10, "Recomendaciones Estratégicas", ln=True)
            pdf.ln(10)
            
            recommendations = [
                "Reforzar controles para detectar Aflatoxinas, especialmente en productos de origen indio y estadounidense.",
                "Incrementar muestreo para Listeria monocytogenes en lácteos, con énfasis en productos de países con mayor incidencia.",
                "Mejorar sistemas de etiquetado para reducir alertas por alérgenos no declarados, implementando controles más rigurosos.",
                "Establecer vigilancia reforzada para productos provenientes de los cinco países con mayor incidencia de alertas.",
                "Implementar programa de seguimiento especial para categorías de alto riesgo, con foco en granos y productos lácteos."
            ]
            
            pdf.set_font('Arial', '', 12)
            for i, rec in enumerate(recommendations, 1):
                pdf.multi_cell(0, 8, f"{i}. {rec}")
                pdf.ln(5)
            
            # Página de conclusiones
            pdf.add_page()
            pdf.set_font('Arial', 'B', 18)
            pdf.cell(0, 10, "Conclusiones y Valor del Sistema", ln=True)
            pdf.ln(10)
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, "Valor estratégico del sistema:", ln=True)
            pdf.ln(5)
            
            value_points = [
                "Transformación de miles de alertas en información accionable para toma de decisiones estratégicas y operativas.",
                "Identificación objetiva de prioridades basada en análisis cuantitativo de riesgos, permitiendo una asignación eficiente de recursos.",
                "Detección de patrones geográficos, temporales y por categoría no evidentes en análisis manual de datos.",
                "Comunicación clara de riesgos complejos mediante visualizaciones estratégicas adaptadas a diferentes audiencias.",
                "Base para decisiones proactivas de control, inspección y aprovisionamiento, reduciendo costos y mejorando la seguridad."
            ]
            
            pdf.set_font('Arial', '', 12)
            for i, point in enumerate(value_points, 1):
                pdf.multi_cell(0, 8, f"{i}. {point}")
                pdf.ln(4)
            
            pdf.ln(10)
            pdf.set_font('Arial', 'I', 12)
            pdf.multi_cell(0, 8, "Este sistema de análisis proporciona un valor significativo para autoridades reguladoras, empresas alimentarias, aseguradoras y consumidores, al traducir datos complejos en información estratégica para la gestión efectiva de riesgos.")
            
            # Guardar el PDF
            pdf_path = os.path.join(self.output_dir, 'executive', f'resumen_ejecutivo_{self.report_date}.pdf')
            pdf.output(pdf_path)
            
            # Limpiar archivos temporales
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
            
            logger.info(f"Presentación ejecutiva generada: {pdf_path}")
            return pdf_path
        
        except Exception as e:
            logger.error(f"Error al generar presentación ejecutiva: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def _generate_presentation_figures(self, temp_dir):
        """Genera figuras para la presentación ejecutiva"""
        # Configurar estilo para gráficos de presentación
        plt.style.use('ggplot')
        plt.rcParams['figure.figsize'] = (12, 7)
        plt.rcParams['font.size'] = 14
        
        # 1. Distribución por nivel de riesgo (pie chart)
        risk_dist = self.df['nivel_riesgo'].value_counts()
        
        plt.figure(figsize=(10, 7))
        colors = ['#ff9999', '#ffcc99', '#99cc99']
        plt.pie(risk_dist, labels=risk_dist.index, autopct='%1.1f%%', startangle=90, colors=colors)
        plt.axis('equal')
        plt.title('Distribución de Alertas por Nivel de Riesgo', fontsize=16)
        plt.tight_layout()
        plt.savefig(os.path.join(temp_dir, 'risk_distribution.png'), dpi=150, bbox_inches='tight')
        plt.close()
        
        # 2. Distribución de nivel de riesgo por categoría
        risk_by_category = pd.crosstab(self.df['category'], self.df['nivel_riesgo'])
        
        plt.figure(figsize=(14, 8))
        risk_by_category.plot(kind='bar', stacked=True, color=colors)
        plt.title('Distribución de Nivel de Riesgo por Categoría de Producto', fontsize=16)
        plt.xlabel('Categoría de Producto', fontsize=14)
        plt.ylabel('Número de Alertas', fontsize=14)
        plt.legend(title='Nivel de Riesgo')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(os.path.join(temp_dir, 'risk_by_category.png'), dpi=150, bbox_inches='tight')
        plt.close()
        
        # 3. Top países de origen
        top_countries = self.df['country_origin'].value_counts().head(10)
        
        plt.figure(figsize=(14, 8))
        top_countries.plot(kind='bar', color='#8c564b')
        plt.title('Top 10 Países de Origen con Más Alertas', fontsize=16)
        plt.xlabel('País de Origen', fontsize=14)
        plt.ylabel('Número de Alertas', fontsize=14)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(os.path.join(temp_dir, 'top_countries.png'), dpi=150, bbox_inches='tight')
        plt.close()
        
        # 4. Tendencia temporal de alertas por nivel de riesgo
        df_year_risk = self.df[self.df['year'].notna()].copy()
        year_risk_counts = pd.crosstab(df_year_risk['year'], df_year_risk['nivel_riesgo'])
        
        plt.figure(figsize=(14, 8))
        year_risk_counts.plot(kind='line', marker='o', linewidth=2, color=['#d62728', '#ff7f0e', '#2ca02c'])
        plt.title('Evolución Anual de Alertas por Nivel de Riesgo', fontsize=16)
        plt.xlabel('Año', fontsize=14)
        plt.ylabel('Número de Alertas', fontsize=14)
        plt.legend(title='Nivel de Riesgo')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        plt.savefig(os.path.join(temp_dir, 'risk_trends.png'), dpi=150, bbox_inches='tight')
        plt.close()
        
        # 5. Top tipos de peligros
        hazard_counts = self.df['hazard_type'].dropna().value_counts().head(10)
        
        plt.figure(figsize=(14, 8))
        ax = hazard_counts.plot(kind='bar', color=plt.cm.tab10.colors)
        plt.title('Top 10 Tipos de Peligros Más Frecuentes', fontsize=16)
        plt.xlabel('Tipo de Peligro', fontsize=14)
        plt.ylabel('Número de Alertas', fontsize=14)
        
        # Truncar etiquetas largas
        labels = [label.get_text() for label in ax.get_xticklabels()]
        short_labels = [label[:30] + '...' if len(label) > 30 else label for label in labels]
        ax.set_xticklabels(short_labels, rotation=45, ha='right')
        
        plt.tight_layout()
        plt.savefig(os.path.join(temp_dir, 'top_hazards.png'), dpi=150, bbox_inches='tight')
        plt.close()

    def generate_report(self, report_type='all'):
        """
        Interfaz principal para generar todos o un tipo específico de informe
        
        Args:
            report_type (str): Tipo de informe a generar ('all', 'excel', 'pdf', 'executive')
            
        Returns:
            dict/str: Rutas a los informes generados o None si hubo un error
        """
        # Cargar y procesar datos
        if not self.load_and_process_data():
            logger.error("Error al cargar y procesar datos para generación de informes")
            return None
        
        # Generar informes según lo solicitado
        if report_type == 'all':
            excel_path = self.generate_excel_report()
            pdf_path = self.generate_notebook_pdf()
            executive_path = self.generate_executive_presentation()
            
            results = {
                'excel': excel_path,
                'notebook_pdf': pdf_path,
                'executive': executive_path
            }
            
            # Log de resultados
            logger.info("\nResumen de informes generados:")
            logger.info(f"1. Informe Excel por categorías: {excel_path}")
            logger.info(f"2. PDF del notebook completo: {pdf_path}")
            logger.info(f"3. Presentación ejecutiva: {executive_path}")
            
            # Devolver rutas solo si todos los informes se generaron correctamente
            # Modificado para devolver los resultados incluso si algunos fallaron
            return results
            
        elif report_type == 'excel':
            return self.generate_excel_report()
        elif report_type == 'pdf':
            return self.generate_notebook_pdf()
        elif report_type == 'executive':
            return self.generate_executive_presentation()
        else:
            logger.error(f"Tipo de informe no válido: {report_type}")
            return None

def main():
    """Función principal para ejecutar desde línea de comandos"""
    parser = argparse.ArgumentParser(description='Generador de Informes de Riesgo Alimentario')
    parser.add_argument('--data', type=str, help='Ruta al archivo CSV de datos consolidados')
    parser.add_argument('--output', type=str, help='Directorio donde se guardarán los informes')
    parser.add_argument('--notebook', type=str, help='Ruta al notebook de análisis')
    parser.add_argument('--type', type=str, choices=['all', 'excel', 'pdf', 'executive'],
                       default='all', help='Tipo de informe a generar')
    
    args = parser.parse_args()
    
    # Configurar logging para uso como script independiente
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(f"logs/report_generator_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
            logging.StreamHandler()
        ]
    )
    
    # Crear generador de informes con argumentos opcionales
    report_gen = AlertReportGenerator(
        data_path=args.data,
        output_dir=args.output,
        notebook_path=args.notebook
    )
    
    # Generar informe del tipo especificado
    result = report_gen.generate_report(report_type=args.type)
    
    if result:
        print("\nInformes generados exitosamente.")
        if isinstance(result, dict):
            for report_type, path in result.items():
                if path:
                    print(f"- {report_type}: {path}")
    else:
        print("\nError al generar informes. Revise los logs para más detalles.")

if __name__ == "__main__":
    main()